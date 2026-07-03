#!/usr/bin/env python3
"""
Build two detailed, per-feature audit tracker workbooks covering the last 36h of work:

  Workbook 1 — Night 1 (2026-06-20): E2E tab audit (156 findings, all merged to main)
  Workbook 2 — Night 2 (2026-06-21): Premium scrub (450 findings) + CoachHelm engine audit (25)

Each workbook: a Summary index, an "All Findings" flat sheet, then ONE sheet per feature/tab.
Columns: ID / Severity / Kind / Original Finding / Fix (what it does) / Status (Fixed when fixed)
/ Confirmed Fixed (before -> after) / File / PR-Wave.

Deterministic — no agents, no API. Run: python3 scripts/build_audit_tracker.py
"""
import csv, json, re, os
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = "/Users/ricknini/Downloads/helmv3"
N1_CSV = f"{ROOT}/docs/archive/2026-06/audits/_e2e_tab_audit_2026-06-20/REMEDIATION_LEDGER.csv"
N2_CSV = f"{ROOT}/docs/archive/2026-06/audits/_premium_scrub_2026-06-21/FINDINGS_CALIBRATED.csv"
CH_MD  = f"{ROOT}/docs/audits/COACHHELM_MASTER_ENGINE_FEATURE_REMEDIATION_AUDIT_2026-06-21.md"
PHASE2 = "/tmp/phase2_outcomes.json"

OUT1 = f"{ROOT}/docs/audits/HELM_AUDIT_TRACKER_Night1_E2E_2026-06-20.xlsx"
OUT2 = f"{ROOT}/docs/audits/HELM_AUDIT_TRACKER_Night2_Premium-CoachHelm_2026-06-21.xlsx"

# ───────────────────────── palette ─────────────────────────
GREEN   = "16A34A"; GREEN_D = "0F7A37"; GREEN_DK = "0B5D2A"
CREAM   = "FFFEFA"; INK = "1C1917"; WHITE = "FFFFFF"
SEV_FILL = {"Critical":"FCA5A5","High":"FDBA74","Medium":"FDE68A","Low":"E2E8F0"}
SEV_TX   = {"Critical":"7F1D1D","High":"7C2D12","Medium":"713F12","Low":"334155"}
ST_FILL  = {"Fixed":"BBF7D0","Pending":"F1F5F9","Skipped":"E0E7FF","Deferred":"FEF3C7","Progress":"BFDBFE"}
ST_TX    = {"Fixed":"14532D","Pending":"475569","Skipped":"3730A3","Deferred":"92400E","Progress":"1E40AF"}

thin = Side(style="thin", color="E5E7EB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HFONT  = Font(name="Calibri", bold=True, color=WHITE, size=10)
HFILL  = PatternFill("solid", fgColor=GREEN_DK)
CFONT  = Font(name="Calibri", size=9, color=INK)
WRAP   = Alignment(wrap_text=True, vertical="top")
CTR    = Alignment(horizontal="center", vertical="top", wrap_text=True)

# ───────────────────────── helpers ─────────────────────────
def norm_sev(s):
    s = (s or "").strip().lower()
    return {"critical":"Critical","high":"High","medium":"Medium","med":"Medium","low":"Low"}.get(s, s.title() or "—")

def classify(text):
    t = (text or "").lower()
    pairs = [
        ("Security / RLS", ["rls","is_public","idor","anon ","leak","policy","grant","unauthor","permission"]),
        ("Data / Schema",  ["column","42703","information_schema","not null","schema","migration","does not exist","persist","upsert","jsonb"]),
        ("Engine / Logic", ["attribution","prediction","generator","composite","orchestrator","weight","causality","validate","rollup","diagnosis","confidence"]),
        ("Accessibility",  ["focus ring","contrast","aria","wcag","keyboard","a11y","label","screen reader","htmlfor","4.5:1","3:1"]),
        ("States",         ["skeleton","empty state","loading","error state","masked as empty","no rounds","cheerful-empty","shape-match"]),
        ("Wiring",         ["deep-link","dead ","no-op","not wired","silently drop","onclick","handler","unwired","orphan","never fires","does nothing"]),
        ("UX Feedback",    ["toast","optimistic","undo","revert","feedback","flicker","silent"]),
        ("Navigation / IA",["breadcrumb","nav ","sidebar","route","duplicate surface","ia "]),
    ]
    for label, kws in pairs:
        if any(k in t for k in kws):
            return label
    return "UI / UX"

def est_height(cells, widths):
    """Rough auto-height: longest wrapped column drives the row height."""
    lines = 1
    for val, w in zip(cells, widths):
        s = str(val or "")
        cpl = max(8, int(w * 1.05))
        rows = sum(max(1, (len(seg)//cpl)+1) for seg in s.split("\n"))
        lines = max(lines, rows)
    return min(420, 15 * min(lines, 28) + 4)

def sheet_name(name, used):
    n = name.replace("holistic:global-", "Global · ").replace("holistic:", "Global · ")
    n = n.replace("/", "-").replace(":", " · ")
    bad = '[]*?\\'
    n = "".join(c for c in n if c not in bad).strip()
    n = n.title() if n.startswith("Global · ") else n
    n = n[:31] or "Sheet"
    base, i = n, 2
    while n.lower() in used:
        suf = f" ({i})"; n = base[:31-len(suf)] + suf; i += 1
    used.add(n.lower()); return n

SEV_ORDER = {"Critical":0,"High":1,"Medium":2,"Low":3}

def style_header(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h); cell.font = HFONT; cell.fill = HFILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

def write_rows(ws, rows, widths, sev_col, status_col):
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val); cell.font = CFONT; cell.border = BORDER
            cell.alignment = CTR if c in (sev_col, status_col, len(row)-1+1 if False else -99) else WRAP
        # severity color (match leading token; calibrated cells read "Medium\n(was Critical)")
        sv = str(row[sev_col-1]).split("\n")[0].strip()
        if sv in SEV_FILL:
            sc = ws.cell(r, sev_col); sc.fill = PatternFill("solid", fgColor=SEV_FILL[sv])
            sc.font = Font(name="Calibri", size=9, bold=True, color=SEV_TX[sv]); sc.alignment = CTR
        # status color
        sval = str(row[status_col-1])
        key = ("Fixed" if "Fixed" in sval else "Progress" if "Progress" in sval
               else "Skipped" if "Skip" in sval else "Deferred" if "Defer" in sval else "Pending")
        tc = ws.cell(r, status_col); tc.fill = PatternFill("solid", fgColor=ST_FILL[key])
        tc.font = Font(name="Calibri", size=9, bold=True, color=ST_TX[key]); tc.alignment = CTR
        ws.row_dimensions[r].height = est_height(row, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{len(rows)+1}"

# ═══════════════════════ confirmed-fixed notes (harvested) ═══════════════════════
# Premium Phase-1 criticals (from workflow wi5nmt2lz + error-mask commit e322b4b3)
CONFIRMED_P1 = {
"P002":"BEFORE: dashboard/page.tsx caught the coach payload fetch failure and rendered zeroed emptyData (rosterSize 0, 'No rounds logged yet') — an outage looked identical to a healthy empty team AND bypassed the already-wired retry boundary. AFTER: the try/catch is removed so a real DB/network outage throws to the route error.tsx (RouteErrorBoundary, offers retry); a genuinely new team still returns empty arrays/zero counts WITHOUT throwing.",
"P426":"BEFORE: dashboard home fabricated zeroed stats on a DB failure (indistinguishable from an empty team). AFTER: same fix as P002 — the outage surfaces via error.tsx retry instead of masquerading as empty.",
"P425":"BEFORE: rounds/page.tsx left the rounds list empty on a fetch error — 'the query failed' was indistinguishable from 'no rounds'. AFTER: both the coach and player branches check fetchAllRowsResult().error (and the in-progress query error) and throw, so a real failure hits error.tsx's retry; an empty list now only ever means 'no rounds'.",
"P017":"BEFORE: in team-category-insights.ts a membersError (real fetch failure) returned the same neutral empty-team data as a genuinely empty team. AFTER: membersError now returns { success:false } so a fetch failure is distinguishable from an empty team (calibrated down from Critical to Medium, but fixed in the Phase-1 pass).",
"P386":"BEFORE: whats-new/loading.tsx only had the legacy glass+warm skeleton, so the flag-on Fairway load was a chrome swap, not a fade. AFTER: loading.tsx is redesign-aware — flag-on renders a Fairway skeleton matching FairwayWhatsNew (max-w-[720px], ViewHeader-shaped block, matte Surface day-groups, bg-canvas, token-only) + role=status/aria-busy. Legacy skeleton preserved.",
"P296":"BEFORE: golf_document_versions.file_url is text NOT NULL, but uploadNewVersion/createDocument/revertToVersion all omitted it -> not-null violation on every version insert ('Upload new version' broke). AFTER: file_url added to all three inserts, derived via getPublicUrl(storagePath) like the already-correct createGolfDocument (no new storage calls).",
"P309":"BEFORE: ExpenseForm receipt file picker had zero callers — attached receipts were silently dropped on save. AFTER: handleSubmit now uploads via uploadExpenseReceipt(), uses the returned URL as receipt_url, aborts with an inline error on failure (no silent loss), and shows an 'Uploading receipt…' busy state.",
"P358":"BEFORE: handleCompleteTask silently reverted the optimistic task-complete on failure with zero feedback. AFTER: on !result.success it calls fairwayToast.error(...) so the revert is explained (gate B3 / Nielsen #1, #9).",
"P372":"BEFORE: Settings FieldLabel was a bare <span> (no htmlFor) and inputs rendered bare — ~14 inputs had zero label↔control association. AFTER: a LabeledField wrapper on Base UI's Field primitive auto-wires htmlFor for every text Input + both Selects; button-group/display-only labels stay visual spans. Satisfies WCAG 1.3.1/3.3.2/4.1.2.",
"P373":"BEFORE: Reset-defaults/Mute fanned out ~27 parallel writes each reading the same stale prefs JSONB then writing the whole blob -> last-write-wins clobbering. AFTER: one atomic setAllChannels() server action (single read-modify-write) called once by both bulk handlers, with optimistic update + rollback + toast + a bulk-pending button guard.",
"P406":"BEFORE: Command Palette 'Players' row linked to /roster?playerId= (the param was silently dropped — roster has no searchParams). AFTER: repointed to /golf/dashboard/players/${id}, the real per-player surface the roster card's own CTA uses.",
"P435":"Duplicate of P406 (same Players control) — fixed once via the same repoint.",
"P407":"BEFORE: palette 'Recent insights' ?id= deep-link was ignored. AFTER: FairwayCoachHelmSignals seeds openRowId from sp.id so the targeted insight panel opens once rows load, + a one-shot effect strips ?id= from the URL; added id to InsightsPageProps.",
"P436":"Duplicate of P407 (same Recent-insights control) — fixed in the live Fairway path (FairwayCoachHelmSignals), same change set.",
"P415":"BEFORE: cream text on bg-accent-500 measured 3.01:1 — below WCAG 4.5:1 for normal text. AFTER: primary Button + IconButton fills darkened to accent-700 rest (5.86:1) / accent-800 hover (8.10:1) using existing tokens, keeping cream text; accent-500 retained as focus-ring color.",
"P416":"BEFORE: form focus ring accent-500/70 over cream measured 1.97:1 — below the 3:1 focus-indicator floor (SC 1.4.11/2.4.7). AFTER: solid ring-accent-600 in fieldControlBase (≥3.7:1 on every Fairway surface), token-consistent.",
"P417":"BEFORE: the bare editorial event-title input had border-none/outline-none with NO focus cue (WCAG 2.4.7 break). AFTER: focus-within accent ring on its wrapper, matching the Fairway Input pattern; one line, type-clean.",
}
# CoachHelm P0 criticals (from workflow w06u31hmr)
CONFIRMED_CH = {
"P0-01":"BEFORE: computeAttribution fed a direction-agnostic delta into nextWeight, so a drop in a lower-is-better metric (putts/penalties/score-to-par) was learned as a REGRESSION and could lower the weight. AFTER: a canonical METRIC_DIRECTION source of truth + improvementSign; attribution stores raw_delta (DB delta) AND improvement_lift (direction-corrected, DB lift) and feeds ONLY improvement_lift into nextWeight — a +6→+3 score-to-par drop now RAISES the weight; identical real improvements in higher/lower-better metrics give the same positive lift. Tests added. FOLLOW-UP: one-time rebuild of golf_coachhelm_coach_weights from corrected history (data backfill, not done).",
"P0-02":"BEFORE: predictPerformance() was called with no args, so due_date defaulted to today → same-day window → never validatable (779/783 same-day-due in prod); the validator averaged an often-inverted [created,due] window that could even grade against a round played BEFORE the prediction. AFTER: due_date = created_at + 14d (resolveDueDate; honors genuine future events); validation grades the FIRST completed round strictly after created_at through end-of-due-date and records related_round_id; same-day/inverted-horizon predictions are retired (error_category='invalid_horizon') and excluded from rollups. 51 tests added.",
"P0-03":"BEFORE: compose() returned unverified LLM text (used_llm=true, citations_verified=false) and both player cards ignored citations_verified — swapping fabricated numbers in as fact. AFTER: compose() retries once with unmatched-token feedback; if still unverified it DISCARDS the LLM text and returns the deterministic fallback (used_llm=false), logging reason='verification_failed' with the offending tokens. RoundReviewLlmCard + HeroNarrativeCard now only swap in LLM prose / show the AI badge when ok && used_llm && citations_verified. Tests added (+ inverse unverified→fallback test).",
"P0-04":"BEFORE: BaseGenerator.run() caught its own throw and resolved with { id:null, gated:false } — identical to a clean no-data exit — so the orchestrator's Promise.allSettled counted the crash as a SUCCESS and the round was marked fully analyzed. AFTER: every run() exit path stamps a typed RunStatus (generated/gated/no_data/standing_lag/failed); the orchestrator routes a resolved status==='failed' into failures, threading through insights.ts → post-round-trigger to mark the round PARTIAL (coachhelm_failure_reason='engine_partial_failure'). Gated/no-data/standing-lag never count as failed. Tests added.",
"P0-05":"BEFORE: 322 V3 rows had the diagnosis only in prose content — 0 had evidence.root_cause / drivers / confidence_reason, so downstream surfaces couldn't filter/audit/learn. AFTER: a typed Diagnosis (symptom, root_cause, causality_level, drivers[], recommended_action, confidence_reason); generator-base now synthesizes a machine-readable diagnosis for EVERY emitted V3 row, drivers cite the measured metric/value/sample (no fabrication), and causality_level is pinned to 'inferred_hypothesis' so a single-metric verdict renders as a coach hypothesis, not fact. Tests added.",
"P0-06":"BEFORE: composite rules stamped static confidence factors (sample_adequacy 0.8 / recency 1.0 / variance 0.5) and bypassed BaseGenerator.run(), so upsert blended them to ~0.77 with no factors_measured flag, empty windows, no causality label (0/26 rows measured). AFTER: synthesis.normalizeCompositeEvidence() recomputes sample_adequacy honestly from the real combined sample_n (target 30), forces factors_measured=false, defaults causality to inferred_hypothesis, CAPS inferred chains at 0.6 (baked into sample_adequacy so it survives the upsert recompute), and copies the widest source window into empty windows; observed-sequence chains are NOT capped. Rules now label estimated rates 'Estimated'. 10 tests added.",
}
phase2 = json.load(open(PHASE2)) if os.path.exists(PHASE2) else {}
# Live status from the running Phase 3/4 workflow (harvested from transcripts)
LIVE = json.load(open("/tmp/live_outcomes.json")) if os.path.exists("/tmp/live_outcomes.json") else {}
SCOPE = set(json.load(open("/tmp/run_scope.json"))) if os.path.exists("/tmp/run_scope.json") else set()

def live_status(rid, stage_prefix="Phase 3/4"):
    """Return (status, stage, confirmed) if rid is live/in-scope, else None."""
    if rid in LIVE:
        act = LIVE[rid]["action"]; note = LIVE[rid].get("note","")
        if act in ("fixed","already-ok"):
            pre = "ALREADY CORRECT (verified): " if act=="already-ok" else ""
            return ("✅ Fixed", f"{stage_prefix} · committed 1402827c (branch fix/premium-remediation)", pre+note)
        if act == "demoted":
            return ("🟡 Deferred (multi-file refactor)", "Deferred — needs dedicated build", "DEFERRED: "+note)
        if act == "skipped-legacy":
            return ("⏭️ Skipped (legacy path)", "Skipped — legacy, out of scope", "SKIPPED: "+note)
    if rid in SCOPE:
        return ("🔵 In Progress", f"{stage_prefix} · agent running now", "")
    return None

# ═══════════════════════ load Night 1 ═══════════════════════
def load_night1():
    rows = list(csv.DictReader(open(N1_CSV)))
    out = []
    for r in rows:
        st = r["status"].strip()
        status = {"done":"✅ Fixed","db-done":"✅ Fixed","skip-dormant":"⏭️ Skipped (dormant)"}.get(st, st)
        conf = r["verification"].strip() if st in ("done","db-done") else (
               "Dormant/legacy feature — intentionally not fixed this pass." if st=="skip-dormant" else "")
        pr = r["pr"].strip()
        prw = (f"{pr} (merged to main)" if pr.startswith("#") else (pr if pr else "—"))
        out.append({
            "feature": r["tab"].strip(),
            "id": r["id"], "sev": norm_sev(r["severity"]),
            "kind": classify(r["issue"]), "sub": r["unit"].strip(),
            "finding": r["issue"].strip().replace("`",""),
            "fix": r["cluster"].strip(),
            "status": status, "confirmed": conf,
            "file": r["file"].strip().replace("`",""), "pr": prw,
        })
    return out

# ═══════════════════════ load Night 2 premium ═══════════════════════
def load_premium():
    rows = list(csv.DictReader(open(N2_CSV)))
    out = []
    for r in rows:
        rid = r["id"]; sev = norm_sev(r["severity"]); orig = norm_sev(r["orig_severity"])
        # status — live run (Phase 3/4) takes precedence, then Phase 1, then Phase 2
        ls = live_status(rid)
        if ls and rid not in CONFIRMED_P1:
            status, stage, conf = ls
        elif rid in CONFIRMED_P1:
            status, stage, conf = "✅ Fixed", "Phase 1 · branch fix/premium-remediation (committed)", CONFIRMED_P1[rid]
        elif rid in phase2:
            act, note = phase2[rid]
            if act in ("fixed","already-ok"):
                status = "✅ Fixed"; stage = "Phase 2 · working tree (pre-commit)"
                conf = ("ALREADY CORRECT (verified): " if act=="already-ok" else "") + note
            elif act == "demoted":
                status = "🟡 Deferred (multi-file refactor)"; stage = "Deferred — needs dedicated build"; conf = "DEFERRED: " + note
            elif act == "skipped-legacy":
                status = "⏭️ Skipped (legacy path)"; stage = "Skipped — legacy, out of scope"; conf = "SKIPPED: " + note
            else:
                status, stage, conf = "🟡 Pending", "—", note
        else:
            status, stage, conf = "🟡 Pending", "—", ""
        cal = ""
        if (r.get("calibrated") or "").strip()=="yes" and orig != sev:
            cal = f"  (orig: {orig})"
        out.append({
            "feature": r["feature"].strip(),
            "id": rid, "sev": sev, "orig": orig,
            "kind": (r["dimension"].strip() or classify(r["evidence"])).replace("-", " ").title(),
            "sub": r["sub_feature"].strip(),
            "finding": (r["evidence"].strip() or r["target_gap"].strip()),
            "fix": r["recommendation"].strip(),
            "status": status, "confirmed": conf,
            "file": r["file"].strip(), "pr": stage,
        })
    return out

# ═══════════════════════ load CoachHelm md ═══════════════════════
def load_coachhelm():
    txt = open(CH_MD).read()
    # split into findings on '### Pn-nn. Title'
    parts = re.split(r'\n### (P\d-\d+)\. ', txt)
    out = []
    sevmap = {"P0":"Critical","P1":"High","P2":"Medium","P3":"Low"}
    for i in range(1, len(parts), 2):
        rid = parts[i]; body = parts[i+1]
        title = body.split("\n",1)[0].strip()
        def grab(label):
            m = re.search(rf'\*\*{label}:?\*\*\s*(.*?)(?=\n\*\*|\Z)', body, re.S)
            return (m.group(1).strip() if m else "")
        why = grab("Why this matters")
        fix = grab("Fix")
        # condense fix bullets
        fix_c = re.sub(r'\n+', " ", re.sub(r'```json.*?```', "[structured JSON schema]", fix, flags=re.S))
        fix_c = re.sub(r'\s*[-•]\s*', " • ", fix_c).strip(" •")
        # file from evidence
        ev = grab("Evidence")
        fm = re.search(r'`(src/[^`]+)`', ev)
        sev = sevmap[rid[:2]]
        if rid in CONFIRMED_CH:
            status, stage, conf = "✅ Fixed", "Phase 1 (CoachHelm) · branch (committed a0dbbeb2)", CONFIRMED_CH[rid]
        else:
            ls = live_status(rid, "Phase 4 (CoachHelm)")
            status, stage, conf = ls if ls else ("🟡 Pending", "—", "")
        out.append({
            "feature": "CoachHelm Engine Audit",
            "id": rid, "sev": sev, "orig": sev,
            "kind": classify(why + " " + fix), "sub": title,
            "finding": why, "fix": fix_c,
            "status": status, "confirmed": conf,
            "file": (fm.group(1) if fm else "see audit doc"), "pr": stage,
        })
    return out

# ═══════════════════════ workbook builders ═══════════════════════
def add_summary(wb, title, subtitle, feats, totals):
    ws = wb.active; ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws["A1"] = title; ws["A1"].font = Font(bold=True, size=18, color=GREEN_DK)
    ws["A2"] = subtitle; ws["A2"].font = Font(size=10, color="64748B", italic=True)
    # totals strip
    r0 = 4
    cards = [("Total", totals["total"], "475569"), ("Critical", totals.get("Critical",0), SEV_TX["Critical"]),
             ("High", totals.get("High",0), SEV_TX["High"]), ("Medium", totals.get("Medium",0), SEV_TX["Medium"]),
             ("Low", totals.get("Low",0), SEV_TX["Low"]), ("Fixed", totals["fixed"], ST_TX["Fixed"]),
             ("In Progress", totals.get("progress",0), ST_TX["Progress"]),
             ("Pending", totals["pending"], ST_TX["Pending"]), ("Skipped/Def.", totals["skipped"], ST_TX["Skipped"])]
    for c,(lab,val,col) in enumerate(cards,1):
        ws.cell(r0,c,lab).font = Font(bold=True, size=9, color="FFFFFF")
        ws.cell(r0,c).fill = PatternFill("solid", fgColor=GREEN_DK)
        ws.cell(r0,c).alignment = Alignment(horizontal="center")
        v = ws.cell(r0+1,c,val); v.font = Font(bold=True, size=16, color=col)
        v.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = 13
    # per-feature table
    hr = r0+3
    heads = ["Feature / Tab","Total","Critical","High","Medium","Low","Fixed","Pending","Skip/Defer","% Fixed"]
    for c,h in enumerate(heads,1):
        cell = ws.cell(hr,c,h); cell.font=HFONT; cell.fill=HFILL; cell.border=BORDER
        cell.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
    ws.column_dimensions["A"].width = 34
    for nm in [b for b in heads[1:]]:
        pass
    rr = hr+1
    for f in feats:
        pct = int(round(100*f["fixed"]/f["total"])) if f["total"] else 0
        vals = [f["name"], f["total"], f.get("Critical",0), f.get("High",0), f.get("Medium",0), f.get("Low",0),
                f["fixed"], f["pending"], f["skipped"], f"{pct}%"]
        for c,v in enumerate(vals,1):
            cell = ws.cell(rr,c,v); cell.font=CFONT; cell.border=BORDER
            cell.alignment = Alignment(vertical="center") if c==1 else Alignment(horizontal="center",vertical="center")
        # tint % fixed
        pcell = ws.cell(rr,10);
        pcell.fill = PatternFill("solid", fgColor=("BBF7D0" if pct==100 else ("FDE68A" if pct>0 else "F1F5F9")))
        pcell.font = Font(bold=True, size=9, color=("14532D" if pct==100 else "713F12"))
        rr += 1
    ws.freeze_panes = f"A{hr+1}"
    ws.auto_filter.ref = f"A{hr}:J{rr-1}"

def build_workbook(path, title, subtitle, data, with_orig):
    wb = Workbook()
    # group by feature
    groups = defaultdict(list)
    for d in data: groups[d["feature"]].append(d)
    # feature summary rows (sorted: most criticals/highs first, then name)
    feat_rows = []
    for name, items in groups.items():
        sc = Counter(i["sev"] for i in items)
        fixed = sum(1 for i in items if "Fixed" in i["status"])
        skipped = sum(1 for i in items if ("Skip" in i["status"] or "Defer" in i["status"]))
        progress = sum(1 for i in items if "Progress" in i["status"])
        pending = len(items) - fixed - skipped - progress
        feat_rows.append({"name":name,"total":len(items),"fixed":fixed,"pending":pending,"skipped":skipped,"progress":progress,
                          "Critical":sc.get("Critical",0),"High":sc.get("High",0),"Medium":sc.get("Medium",0),"Low":sc.get("Low",0),
                          "items":items})
    feat_rows.sort(key=lambda f:(-f["Critical"],-f["High"],-f["total"],f["name"]))
    # overall totals
    sc_all = Counter(d["sev"] for d in data)
    tot = {"total":len(data),
           "fixed":sum(1 for d in data if "Fixed" in d["status"]),
           "progress":sum(1 for d in data if "Progress" in d["status"]),
           "skipped":sum(1 for d in data if ("Skip" in d["status"] or "Defer" in d["status"]))}
    tot["pending"] = tot["total"]-tot["fixed"]-tot["skipped"]-tot["progress"]
    tot.update({k:sc_all.get(k,0) for k in ("Critical","High","Medium","Low")})
    add_summary(wb, title, subtitle, feat_rows, tot)

    # column spec
    if with_orig:
        heads = ["ID","Severity","Kind","Sub-Feature","Original Finding","Fix / What It Does",
                 "Status","Confirmed Fixed — Before → After","File","Stage / PR"]
        widths = [9,11,16,24,52,46,20,60,40,30]
        def rowvals(d): return [d["id"],d["sev"]+("\n(was %s)"%d["orig"] if d.get("orig") and d["orig"]!=d["sev"] else ""),
                                d["kind"],d["sub"],d["finding"],d["fix"],d["status"],d["confirmed"],d["file"],d["pr"]]
    else:
        heads = ["ID","Severity","Kind","Sub-Feature","Original Finding","Fix / What It Does",
                 "Status","Confirmed Fixed — Before → After","File","PR / Wave"]
        widths = [9,11,16,22,54,40,18,58,40,26]
        def rowvals(d): return [d["id"],d["sev"],d["kind"],d["sub"],d["finding"],d["fix"],
                                d["status"],d["confirmed"],d["file"],d["pr"]]
    SEVC, STC = 2, 7

    # All-Findings flat sheet
    allws = wb.create_sheet("All Findings")
    style_header(allws, ["Feature"]+heads, [24]+widths)
    flat = []
    for f in feat_rows:
        for d in sorted(f["items"], key=lambda x:(SEV_ORDER.get(x["sev"],9), x["id"])):
            flat.append([f["name"]]+rowvals(d))
    write_rows(allws, flat, [24]+widths, SEVC+1, STC+1)

    # per-feature sheets
    used=set(); used.add("summary"); used.add("all findings")
    for f in feat_rows:
        ws = wb.create_sheet(sheet_name(f["name"], used))
        style_header(ws, heads, widths)
        rows = [rowvals(d) for d in sorted(f["items"], key=lambda x:(SEV_ORDER.get(x["sev"],9), x["id"]))]
        write_rows(ws, rows, widths, SEVC, STC)
    wb.save(path)
    return tot, len(feat_rows)

# ═══════════════════════ run ═══════════════════════
n1 = load_night1()
t1,fc1 = build_workbook(OUT1,
    "Helm Audit Tracker — Night 1 (E2E Tab Audit)",
    "Captured 2026-06-20 · 156 end-to-end findings across every tab · ALL merged to main (PRs #327–#332)",
    n1, with_orig=False)
print(f"WB1 ({fc1} feature tabs): {t1}")

n2 = load_premium() + load_coachhelm()
t2,fc2 = build_workbook(OUT2,
    "Helm Audit Tracker — Night 2 (Premium Scrub + CoachHelm Engine)",
    "Captured 2026-06-21 · 450 premium-scrub findings (calibrated) + 25 CoachHelm engine findings · branch fix/premium-remediation (gated PR, not merged)",
    n2, with_orig=True)
print(f"WB2 ({fc2} feature tabs): {t2}")
print("OUT1:", OUT1)
print("OUT2:", OUT2)
